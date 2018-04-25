import re
from openpyxl import load_workbook, Workbook
from datetime import datetime


def escape_unicode(str):
    if str == '' or str is None:
        return ''

    if(isinstance(str, datetime)):
        return str

    if type(str).__name__ == 'unicode': str = str.encode('utf-8')
    list_pat = ["á|à|ả|ạ|ã|â|ấ|ầ|ẩ|ậ|ẫ|ă|ắ|ằ|ẳ|ặ|ẵ", "Á|À|Ả|Ạ|Ã|Â|Ấ|Ầ|Ẩ|Ậ|Ẫ|Ă|Ắ|Ằ|Ẳ|Ặ|Ẵ",
                "đ", "Đ", "í|ì|ỉ|ị|ĩ", "Í|Ì|Ỉ|Ị|Ĩ", "é|è|ẻ|ẹ|ẽ|ê|ế|ề|ể|ệ|ễ", "É|È|Ẻ|Ẹ|Ẽ|Ê|Ế|Ề|Ể|Ệ|Ễ",
                "ó|ò|ỏ|ọ|õ|ô|ố|ồ|ổ|ộ|ỗ|ơ|ớ|ờ|ở|ợ|ỡ", "Ó|Ò|Ỏ|Ọ|Õ|Ô|Ố|Ồ|Ổ|Ộ|Ỗ|Ơ|Ớ|Ờ|Ở|Ợ|Ỡ",
                "ú|ù|ủ|ụ|ũ|ư|ứ|ừ|ử|ự|ữ", "Ú|Ù|Ủ|Ụ|Ũ|Ư|Ứ|Ừ|Ử|Ự|Ữ", "ý|ỳ|ỷ|ỵ|ỹ", "Ý|Ỳ|Ỷ|Ỵ|Ỹ"]
    list_re = ['a', 'A', 'd', 'D', 'i', 'I', 'e', 'E', 'o', 'O', 'u', 'U', 'y', 'Y']
    for i in range(len(list_pat)):
        str = re.sub(list_pat[i], list_re[i], str)

    return str


wb = Workbook()
ws = wb.active

book = load_workbook('./excel/excel_unicode.xlsx')
sheet = book.active

rows = sheet.rows

values = []

for row in rows:

    for cell in row:
        print(cell.row, cell.column)
        escape_value = escape_unicode(cell.value)
        print(escape_value)
        index = "{}{}".format(cell.column, cell.row)
        ws[index].value = escape_value
        ws.column_dimensions[cell.column].width = 20.0

wb.save('./excel/empty_unicode.xlsx')

